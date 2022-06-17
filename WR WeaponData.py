import math
import openpyxl as xl

##########
# Section 0: Data input
##########

# Open excel file
def init():
    global MAINFILE, wb, ref, stats, names, ids
    MAINFILE = '8.0_Stats_v4.0_20220516.xlsx'
    wb = xl.load_workbook(MAINFILE)
    ref = wb['Weapon Attributes']
    stats = wb['Weapons']
    # to search, use name in ref, id in stats
    names = [i.value for i in ref['B']]
    ids = [i.value for i in stats['A']]

# Automatic data input from datasheet
def input_auto(row):
    # input a row of reference values to generate
    w = {}
    w['id'] = row[0]
    w['name'] = row[1]
    w['ammo'] = row[9]
    w['aps'] = row[10]
    w['ppa'] = row[11]
    w['sintv'] = row[12]
    w['ssintv'] = row[13]
    w['rnum'] = row[16]
    w['rintv'] = row[17]
    w['cd'] = row[18]
    w['rthrs'] = row[19]
    
    # find level-wise data by ID
    w['dpart'] = []
    w['dot'] = []
    w['accum'] = []
    for i in stats.iter_rows(min_row = ids.index(w['id'])+1,
                             max_row = ids.index(w['id'])+25, 
                             min_col = 5,
                             max_col = 7,
                             values_only = True):
        w['dpart'].append(i[0])
        w['dot'].append(i[1])
        w['accum'].append(i[2])
    
    # store properties
    type = row[25]
    rtypes = {'Magazine':'',
              'Reloads while firing':'r',
              'Accumulates charges':'c',
              'Unlimited ammo':'u',
              'Reloads while not firing':'rwnf'}
    w['rtype'] = rtypes.get(type, 'Magazine')
    if w['rtype'] == 'c':
        # turn charge-up weapons into single-shot
        # fires all charges as distinct particles
        w['cd'] += w['ammo'] * w['rintv']
        w['ppa'] = w['ammo']
        w['ammo'] = 1
        w['aps'] = 1
        w['sintv'] = 0
        w['ssintv'] = 0
        w['rnum'] = 1
        w['rintv'] = 0
        w['rthrs'] = 0
    elif w['rtype'] == 'u':
        # give unlimited ammo weapons negative ammo
        # to bypass ammo check
        w['ammo'] = -1
        w['aps'] = 1
        # ppa, sintv, ssintv unchanged
        # reload data modified so that cycle dps works
        # permits reload but reload goes the opposite way
        w['rnum'] = -1
        w['rintv'] = w['sintv']
        w['cd'] = 0
        w['rthrs'] = 0
    elif w['rtype'] == 'rwnf':
        # limit reloading to only at empty ammo
        # (for caclulations only)
        w['rthrs'] = 0
    
    type = row[26]
    dtypes = {'Acceleration':'a',
              'Close combat':'c',
              'Distanced combat':'d',
              'Ignores defence systems':'i'}
    w['dtype'] = dtypes.get(type)
    if w['dtype'] == 'a':
        # all accelerated shot intervals are rounded to 0.01
        # this way it corresponds well to the game
        if w['name'] in ('Punisher','Punisher T','Avenger'):
            # 500m MGs
            w['acctime'] = 3
            w['accsintv'] = 0.06
        if w['name'] in ('Molot','Molot T','Tempest'):
            # 800m MGs
            w['acctime'] = 3
            w['accsintv'] = 0.14
        if w['name'] == 'Viper':
            # seems like 2 sec and +25%/30%(12.32s, exact)
            # https://www.youtube.com/watch?v=ECggW_HMsdk
            w['acctime'] = 2
            w['accsintv'] = 0.08
        if w['rtype'] == 'r':
            # HMGs
            # activation time is notably faster
            w['acctime'] = 2
            w['accsintv'] = 0.06
        
    type = row[28]
    etypes = {'Lock-down':'l',
              'DOT':'d',
              'Freeze':'f',
              'BLASTCHARGE':'b'}
    w['etype'] = etypes.get(type)
    # implemented in the simulation
    # will be done after we get info
    # dot is easiest, dotdmg * bpart
    # blast is okay, 25000 * trunc(btime/atime)
    # freeze is hardest... need one run for each level
    
    return w


##########
# Section 1: Simulate the firing process tick-wise
##########

# Length of gametick, seconds
GAMETICK = 0.001

def to_ticks(x):    # convert seconds to gametick
    return int(x / GAMETICK)

def to_seconds(x):  # convert gametick to seconds
    return round(x * GAMETICK, 3)

# Main simulator
def simulate(w, evaltime=30, debug=False):
    # Weapon stats
    max_ammo = w['ammo']  # full ammo
    ammo = max_ammo # current ammo amount
    aps = w['aps']  # ammo used per shot
    ppa = w['ppa']  # particles per ammo
    
    sintv_normal = to_ticks(w['sintv'])  # normal shot interval
    sintv = sintv_normal  # current shot interval
    ssintv = to_ticks(w['ssintv'])  # shot sub-interval
    
    cd = to_ticks(w['cd'])  # weapon cooldown time (grey)
    rintv = to_ticks(w['rintv'])  # reload interval
    rnum = w['rnum']  # amount of ammo reloaded each time
    rthrs = w['rthrs']  # reload threshold
    
    # acceleration mode
    if w['dtype'] == 'a':
        acctime = to_ticks(w['acctime'])
        sintv_acc = to_ticks(w['accsintv'])
    
    # Timers, unit in ticks
    time = 0         # total time
    stimer = 0       # time since last shot
    sstimer = 0      # time since last fire in a shot event
    cdcntdown = 0    # cooldown countdown
    rtimer = 0       # time since last reload
    acctimer = 0     # time of continuous fire
    
    # Result counters
    shot = 0      # total shot events
    part = 0      # total particles fired
    reload = 0    # total reload events
    sintvs = []   # list of actual shot intervals
    ssintvs = []  # list of actual shot sub-intervals
        
    # Counters for ammo per shot issues
    # index of ammo in a shot event
    # begins counting from 0 (0 is first)
    # when ammo_in_shot == aps, the shot event ends
    # for non-burst weapons (aps = 1), 0 means not shot yet, 1 means shot done
    ammo_in_shot = 0
    
    if debug == True:
        print('time ammo re shot stime rtime')
    
    # Main tick cycle
    while time < to_ticks(evaltime):  # run time default 20s
        if cdcntdown == 0:
            # check reload before firing
            if ammo <= rthrs:
                # reload only begins when ammo falls below threshold
                rtimer += 1
                if rtimer >= rintv:
                    if ammo+rnum > max_ammo:
                        # prevent reload overflow
                        reload += max_ammo - ammo
                        ammo = max_ammo
                        rtimer = 0
                    else:
                        reload += rnum
                        ammo += rnum
                        rtimer = 0
            else:
                rtimer = 0
            
            
            # implement acceleration mode
            if w['dtype'] == 'a':
                if acctimer == acctime:
                    sintv = sintv_acc
                if stimer >= to_ticks(1):
                    # if stop firing for 1 second, stop acceleration
                    sintv = sintv_normal
                    acctimer = 0
                acctimer += 1
            
            # attempt to fire
            # to fire, (1) positive ammo AND (2) sintv reached
            # NOTE: for charge-up weapons, fire instantly after cooldown
            # to simulate these weapons, modify them so that they are single-shot
            while ammo != 0:  # not greater, to leave space for unlimited ammo weapons
                if shot == 0 or stimer >= sintv:
                    # first shot does not need to wait!
                    shot += 1
                    sintvs.append(stimer)
                    stimer = 0
                    ammo_in_shot = 0
                elif ammo_in_shot < aps and sstimer >= ssintv:
                    # Conditions explained:
                    # 1) still has ammo to fire in a shot (always not if aps == 0)
                    # 2) the subshot reload is ready (always if ssintv == 0)
                    pass  # allow to fire
                else:  # if not the time to fire, skip this part
                    break  # no fire
                
                part += ppa
                ammo -= 1
                ammo_in_shot += 1
                ssintvs.append(sstimer)
                sstimer = 0
            
            
            if ammo == 0 and cd != 0:
                # if cooldown exists, begin cooldown when ammo is depleted
                # -1 because this tick has already passed by this point
                cdcntdown = cd - 1
        else:
            # if in cooldown, decrease cooldown countdown
            if ssintv == 0.15:
                if stimer > sintv:
                # exclusively, for burst weapons Taran and Redeemer
                # due to an in-game bug (fixed for Zenit and Chimera)
                # cooldown only starts after the final shot has completely ended  
                    cdcntdown -= 1
            else:
                cdcntdown -= 1
        
        stimer += 1
        sstimer += 1
        time += 1
        
        if debug == True:
            print(format(time,'05'), format(ammo,'3'), format(reload,'3'), format(shot,'3'), format(stimer,'04'), format(rtimer,'04'))
    
    
    if debug == True:
        # Print testing lists
        # These should be identical if aps == 1
        print(sintvs)
        print(ssintvs)
    
    return w, ssintvs


##########
# Section 2: Analysis of generated data
##########

def analyze(w, ssintvs):
    # store secondary stats here, initially with id and name for reference
    ww = {'id':w['id'],
          'name':w['name']}
    
    # time values in this function should be in ticks
    # Reload time
    # reload time will equal sintv for unlimited ammo
    rintvs = math.ceil(w['ammo'] / w['rnum'])
    rtime = rintvs * to_ticks(w['rintv']) + to_ticks(w['cd'])
    if w['ssintv'] == 0.15:
        # addressing Taran-Redeemer bug
        rtime += to_ticks(w['sintv']) - (w['aps']-1) * to_ticks(w['ssintv'])
    
    # Burst/Cycle data
    ww['bammo'] = ssintvs.index(max(ssintvs))
    btime = 0
    for i in ssintvs[:ww['bammo']]:
        btime += i
    ww['bpart'] = ww['bammo'] * w['ppa']
    
    # convert to seconds
    ww['ctime'] = to_seconds(btime + rtime)
    ww['btime'] = to_seconds(btime)
    if w['rtype'] != 'u':
        ww['rtime'] = to_seconds(rtime)
    else:
        # zero canonical reload for unlimited ammo
        ww['rtime'] = 0
    
    # level-based data
    pps = w['ppa'] * w['aps']
    ww['dshot'] = []
    ww['bdmg'] = []
    ww['bdps'] = []
    ww['cdps'] = []
    for i in w['dpart']:
        if i == '~':
            ww['dshot'].append('~')
            ww['bdmg'].append('~')
            ww['bdps'].append('~')
            ww['cdps'].append('~')
        else:
            ww['dshot'].append(i * pps)
            i *= ww['bpart']
            ww['bdmg'].append(i)
            if ww['btime'] == 0:
                ww['bdps'].append('~')
            else:
                ww['bdps'].append(round(i / ww['btime']))
            ww['cdps'].append(round(i / ww['ctime']))
    
    return ww


##########
# Section 3: Output data into the datasheet
##########

def output(ww):
    # WARNING: TO OUTPUT, ALWAYS USE A SEPARATE DATASHEET
    # TO PREVENT UNEXPECTED RESULTS FROM BREAKING THE DATA
    
    # input secondary ref data
    r = names.index(ww['name']) +1
    now = ref[r]
    now[14].value = ww['btime']
    now[15].value = ww['bpart']
    now[20].value = ww['rtime']
    
    # find level-wise data by ID
    for i in stats.iter_rows(min_row = ids.index(ww['id'])+1,
                             max_row = ids.index(ww['id'])+25, 
                             min_col = 8,
                             max_col = 12):
        i[0].value = ww['dshot'].pop(0)
        i[1].value = ww['bdmg'].pop(0)
        i[2].value = ww['bdps'].pop(0)
        i[3].value = ww['cdps'].pop(0)
        
        # not implemented yet
        # i[4].value = ww['atime'].pop(0)
        # i[5].value = ww['admg'].pop(0)
    
    # return id and name to show output successful
    return ww['id'], ww['name']


##########
# Main switch, controls complete automation
##########

def main():
    init()
    print('    >>> Process Initiated.')
    
    for wpn in ref.iter_rows(min_row = 2,
                             max_row = 86,
                             values_only = True):
        print(*output(analyze(*simulate(input_auto(wpn)))))
    
    wb.save(MAINFILE)
    print('    >>> Process Complete.')

main()