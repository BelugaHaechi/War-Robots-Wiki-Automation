import openpyxl as xl

MAINFILE = '8.0_Stats_v4.0_20220516.xlsx'
wb = xl.load_workbook(MAINFILE)
stats = wb['Weapons']
ref = wb['Weapon Attributes']
ids = [i.value for i in stats['A']]

wiki = open('MasterEquipment.txt', 'w', encoding='utf-8')

header = '{{#switch: {{{1}}}'

def writeln(*args, sep=' ', end='\n'):
    args = [str(i) for i in args]
    text = ' '.join(args)
    wiki.write(text+end)

def align(num, digits):
    try:
        return format(int(num), f'<{digits},')
    except:
        return '~'

def generate():
    writeln(header)
    # dict of stat_name:index pairs
    stats_map = {'name':1,
                 'id':0,
                 'tier':2,
                 'slot':3,
                 'range':7,
                 'aoe':8,
                 'available':5,
                 'ammo':9,
                 'ammo per shot':10,
                 'particles per ammo':11,
                 'shot interval':12,
                 'shot subinterval':13,
                 'unload':14,
                 'reload':20,
                 'reload interval':17,
                 'cooldown':18,
                 'reload amount':16}
    for wpn in ref.iter_rows(min_row = 2,
                             max_row = 86,
                             max_col = 36,
                             values_only = True):
        # generate fixed data
        id = wpn[0]
        name = wpn[1]
        if wpn[32] == 'Unlimited ammo':
            # modify stats to make sense
            # convert tuple to list first
            wpn = list(wpn)
            wpn[9] = '∞'
            wpn[10] = 1
            wpn[14] = '∞'
        
        writeln('|', name, '= {{Weapon stats')
        for k,v in stats_map.items():
            if wpn[v] != None:
                writeln(f'| {k} = {wpn[v]}')
            else:
                writeln(f'| {k} = ~')
        
        # generate attribute icons
        attr = 1
        for i in range(29,36):
            if wpn[i] != None:
                writeln(f'| attribute{attr} = {wpn[i]}')
                attr += 1
        while attr <= 6:
            # fill in blanks
            writeln(f'| attribute{attr} = ')
            attr += 1
        
        # generate level data
        for i in stats.iter_rows(min_row = ids.index(id)+1,
                                 max_row = ids.index(id)+25, 
                                 min_col = 3,
                                 max_col = 11,
                                 values_only = True):
            lvl = str(i[1])
            mk = i[0]
            mk = '' if mk=='I' else ('mk2' if mk=='II' else 'mk3')
            parttag = format(mk + 'part' + lvl, '<9')
            bursttag = format(mk + 'burst' + lvl, '<10')
            bdpstag = format(mk + 'bdps' + lvl, '<9')
            cdpstag = format(mk + 'cdps' + lvl, '<9')
            writeln('|', parttag, '=', align(i[2],6),
                    '||', bursttag, '=', align(i[6],7),
                    '||', bdpstag, '=', align(i[7],7),
                    '||', cdpstag, '=', align(i[8],''))
        
        writeln('}}')  # end of display template
        print(id, name)  # print progress
    
    writeln('|#default = {{Icon|Lock icon|24px}} Statistics currently unavailable, please contact wiki staff if you can help :)') # default case, a kind notice
    writeln('}}') # don't forget to close the #switch!

generate()