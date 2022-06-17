import openpyxl as xl
from openpyxl.styles import Font

MAINFILE = 'WR/8.0_Stats_v3.0_20220507.xlsx'
wb = xl.load_workbook(MAINFILE)
print(wb.sheetnames)
stats = wb['Robots']
ref = wb['Robot Attributes']
names = [i.value for i in ref['A']]

wiki = open('WR/MasterRobot.txt', 'w', encoding='utf-8')

# first row is titles
# every 25 rows are one robot
# NAME LVL MK  SPD HP UPCOST UPTIME
# str  int str int int int   int
# counting starts from 1

'''
print([i.value for i in ref[1]])
print(stats.max_row)
print(stats[1])
print([type(i.value) for i in stats[2]])
'''


header = '{{#switch: {{{1}}}'

def rename():
    # completed
    for cell in ref['A']:
        a = cell.value
        '''
        if cell.row % 25 == 1:
            cell.font = Font(name = 'Calibri', bold = True)
        '''
        if a[-1]==' ': cell.value = a[:-1]
        '''
        if ')' in name:
            #cell.value = name[name.index('(')+1:-1] + ' ' + name[:name.index(' ')]
            
        #if name[-1]==' ': cell.value = name[:-1]
        '''
    wb.save(MAINFILE)

def writeln(*args, sep=' ', end='\n'):
    args = [str(i) for i in args]
    text = ' '.join(args)
    wiki.write(text+end)

def generate():
    writeln(header)
    
    current = ''    
    for row in stats.iter_rows(min_row=2, max_col=5, values_only=True):
        name = row[0]
        
        # create new case upon getting new name
        if current != name:
            current = name
            get_heading(name)
            
        # create a line of table
        lvl = str(row[1])
        mk = row[2]
        mk = '' if mk=='I' else ('mk2' if mk=='II' else 'mk3')
        try:
            hp = format(int(row[4]), '<9,')
        except:
            hp = '~'
        try:
            spd = int(row[3])
        except:
            spd = '~'
        hptag = format(mk + 'hp' + lvl, '<7')
        spdtag = format(mk + 'speed' + lvl, '<10')
        writeln('|', hptag, '=', hp,\
              '||', spdtag, '=', spd)
        
        # add template ending when mk3 is reached
        if mk == 'mk3':
            writeln('}}')
    
    writeln('') # default case, a kind notice
    writeln('}}') # don't forget to close the #switch!


def get_heading(name):
    #['Name', 'Tier', 'Class', 'Wiki Class', 'Ability', 'Faction', 'Availability', 'Cost', 'Level', 'ID', 'Seq#']
    row = names.index(name)+1 # excel row
    now = [i.value for i in ref[row]]
    writeln('|', name, '= {{Robot stats')
    writeln(f'| name = {name}')
    writeln(f'| id = {now[0]}')
    writeln(f'| ability = {now[5]}')
    writeln(f'| faction = {now[6]}')
    writeln(f'| level = {now[7]}')
    writeln(f'| tier = {now[2]}')
    writeln(f'| class = {now[3]}')
    writeln(f'| wikiclass = {now[4]}')
    writeln(f'| available = {now[7]}')

generate()